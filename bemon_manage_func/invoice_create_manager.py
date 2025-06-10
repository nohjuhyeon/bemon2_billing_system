
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from datetime import datetime, timedelta
import calendar
from PIL import Image as PILImage  # 이미지 크기 계산을 위해 사용

class InvoiceCreateManager:
    def __init__(self,billing_dict):
        self.billing_info_matching_dict = {
        'DOC_NUM':{'row':6,'column':26},
        'START_DATE':{'row':7,'column':26},
        'END_DATE':{'row':7,'column':33},
        'TODAY_DATE':{'row':8,'column':26},
        'INVOICE_DATE':{'row':9,'column':26},
        'CLOUD_NAME':{'row':10,'column':26},
        'USER_NAME':{'row':9,'column':5},
        'CLOUD_USER_ID':{'row':10,'column':5},
        'CLOUD_USER_NUM':{'row':11,'column':5},
        'TOTAL_CLOUD_USE_AMT':{'row':19,'column':19},
        'TOTAL_CLOUD_USER_PAY_AMT':{'row':19,'column':24},
        'TOTAL_CLOUD_NOTES':{'row':19,'column':29},
        'TOTAL_CHARGE_THIRD_PARTY_USE_AMT':{'row':20,'column':19},
        'TOTAL_CHARGE_THIRD_PARTY_USER_PAY_AMT':{'row':20,'column':24},
        'TOTAL_CHARGE_THIRD_PARTY_NOTES':{'row':20,'column':29},
        'TOTAL_CHARGE_MANAGED_SERVICE_USE_AMT':{'row':21,'column':19},
        'TOTAL_CHARGE_MANAGED_SERVICE_USER_PAY_AMT':{'row':21,'column':24},
        'TOTAL_CHARGE_MANAGED_SERVICE_NOTES':{'row':21,'column':29},
        'TOTAL_CHARGE_OTHER_SERVICE_USE_AMT':{'row':22,'column':19},
        'TOTAL_CHARGE_OTHER_SERVICE_USER_PAY_AMT':{'row':22,'column':24},
        'TOTAL_CHARGE_OTHER_SERVICE_NOTES':{'row':22,'column':29},
        'TOTAL_USER_PAY_AMT':{'row':23,'column':24},
        'TOTAL_PAY_NOTES':{'row':23,'column':29},
        'TOTAL_SALES_DISCOUNT_AMT':{'row':24,'column':24},
        'TOTAL_SALES_DISCOUNT_NOTES':{'row':24,'column':29},
        'TOTAL_COIN_USE_AMT':{'row':25,'column':24},
        'TOTAL_COIN_USE_NOTES':{'row':25,'column':29},
        'TOTAL_DISCOUNT_INCLUDE_AMT':{'row':26,'column':24},
        'TOTAL_DISCOUNT_INCLUDE_NOTES':{'row':26,'column':29},
        'TOTAL_VAT_AMT':{'row':27,'column':24},
        'TOTAL_VAT_NOTES':{'row':27,'column':29},
        'TOTAL_VAT_INCLUDE_AMT':{'row':28,'column':24},
        'TOTAL_VAT_INCLUDE_NOTES':{'row':28,'column':29},
        'TOTAL_VAT_INCLUDE_AMT_2':{'row':13,'column':12},
        }
        self.collection_list_dict = {
            "서비스명":{'start_column':1,'end_column':4,'key':'CLASS_NAME'}, 
            "청구항목":{'start_column':5,'end_column':10,'key':'SERVICE_NAME'}, 
            "상품명":{'start_column':11,'end_column':18,'key':'PRODUCT_NAME'}, 
            "정상가":{'start_column':19,'end_column':23,'key':'USE_AMT'}, 
            "제공가":{'start_column':24,'end_column':28,'key':'USER_PAY_AMT'}, 
            "비고":{'start_column':29,'end_column':37,'key':'NOTES'}
            }
        self.collection_middle_sum_dict = {
            "서비스명":{'start_column':1,'end_column':4,'key':'CLASS_NAME'}, 
            "소계":{'start_column':5,'end_column':18,'key':'MIDDLE_SUM'}, 
            "정상가":{'start_column':19,'end_column':23,'key':'TOTAL_SERVICE_USE_AMT'}, 
            "제공가":{'start_column':24,'end_column':28,'key':'TOTAL_SERVICE_USER_PAY_AMT'}, 
            "비고":{'start_column':29,'end_column':37,'key':'NOTES'}
            }
        self.collection_total_sum_dict = {
            "합계":{'start_column':1,'end_column':18,'key':'TOTAL_SUM'}, 
            "정상가":{'start_column':19,'end_column':23,'key':'TOTAL_USE_AMT'}, 
            "제공가":{'start_column':24,'end_column':28,'key':'TOTAL_USER_PAY_AMT'}, 
            "비고":{'start_column':29,'end_column':37,'key':'NOTES'}
            }

        self.billing_dict = billing_dict
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "이용내역서"


        self.template_path = "bemon_manage_func/invoice_template.xlsx"
        self.wb = load_workbook(self.template_path)
        self.ws = self.wb.active


        self.colleciton_current_row = 31  # 데이터가 시작될 행 번호

        # 스타일 설정
        self.header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        self.middle_sum_fill = PatternFill(start_color="E8E8E8",end_color="E8E8E8",fill_type="solid")
        self.border_style = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
        self.center_alignment = Alignment(horizontal="center", vertical="center")
        self.left_alignment = Alignment(horizontal="left", vertical="center")

    def image_size_modify(self,image_path,targe_width):
        img = Image(image_path)
        
        # 원본 이미지 크기 가져오기
        with PILImage.open(image_path) as pil_img:
            original_width, original_height = pil_img.size

        aspect_ratio =  original_height / original_width

        # 높이를 비율에 맞게 자동 조정
        target_height = int(targe_width * aspect_ratio)

        # 이미지 크기 설정
        img.width = targe_width
        img.height = target_height
        return img


    def image_create(self):
        # 워크시트에 이미지 삽입
        belab_image_path = "resources/image/belab_logo.png"  # 삽입할 이미지 파일 경로
        belab_img = self.image_size_modify(belab_image_path,200)
        # # 워크시트에 이미지 삽입
        self.ws.add_image(belab_img, 'B2')

        # 워크시트에 이미지 삽입
        if self.billing_dict['user']['CLOUD_NAME'] == 'KT':
            cloud_image_path = "resources/image/kt_logo.png"  # 삽입할 이미지 파일 경로
            cloud_img = self.image_size_modify(cloud_image_path,200)
            self.ws.add_image(cloud_img, "AC2")
        elif self.billing_dict['user']['CLOUD_NAME'] == 'NAVER':
            cloud_image_path = "resources/image/naver_logo.png"  # 삽입할 이미지 파일 경로
            cloud_img = self.image_size_modify(cloud_image_path,300)
            self.ws.add_image(cloud_img, "Y2")
        elif self.billing_dict['user']['CLOUD_NAME'] == 'NHN':
            cloud_image_path = "resources/image/nhn_logo.png"  # 삽입할 이미지 파일 경로
            cloud_img = self.image_size_modify(cloud_image_path,300)
            # 이미지를 특정 셀에 삽입 (예: B2)
            self.ws.add_image(cloud_img, "Y2")
            


    def invoice_amt_fommat(self,cell,cell_value):
        cell.number_format = '_(₩* #,##0_);_(-₩* #,##0_);_(₩* "-"??_);_(@_)'
        # 음수 값일 경우 텍스트 색상을 빨간색으로 설정
        if cell_value < 0:
            cell.font = Font(color="FF0000")  # 빨간색 텍스트

    def charge_amount_format(self,cell,cell_value):
        cell.number_format = '_(₩* #,##0_);_(-₩* #,##0_);_(₩* "-"??_);_(@_)'
        if cell_value and cell_value < 0:
            cell.font = Font(color="FF0000")  # 빨간색 텍스트


    def cell_format_setting(self,cell,cell_key,cell_value):
        if 'AMT' in cell_key:
            self.charge_amount_format(cell,cell_value)
        if cell_key == 'MIDDLE_SUM' or 'TOTAL_SERVICE' in cell_key:
            cell.alignment = self.center_alignment
            cell.fill = self.middle_sum_fill
            cell.font = Font(bold=True,size=9)        
        if cell_key in ['USER_NAME','CLOUD_USER_ID','CLOUD_USER_NUM']:
            cell.alignment = self.left_alignment
        pass

    def date_format(self,billing_month):
        year = int(billing_month[:4])  # 앞 4자리는 연도
        month = int(billing_month[4:])  # 뒤 2자리는 월

        # 해당 월의 시작일
        start_date = datetime(year=year, month=month, day=1)
        start_date = start_date.strftime("%Y-%m-%d")
        # 해당 월의 마지막 날 계산
        last_day = calendar.monthrange(year, month)[1]  # 해당 월의 마지막 날 (예: 31)
        end_date = datetime(year=year, month=month, day=last_day)
        end_date = end_date.strftime("%Y-%m-%d")

        invoice_date = datetime(year=year, month=month, day = 15)
        invoice_date = invoice_date.strftime("%Y-%m-%d")
        today = datetime.today()
        today_date = today.strftime("%Y-%m-%d")
        return start_date,end_date,invoice_date,today_date

    def collection_data_insert(self,start_column, end_column,cell_key, cell_value):
        self.ws.merge_cells(start_row=self.colleciton_current_row, start_column=start_column, end_row=self.colleciton_current_row, end_column=end_column)
        cell = self.ws.cell(row=self.colleciton_current_row, column=start_column, value=cell_value)
        cell.font = Font(size=9)
        self.cell_format_setting(cell,cell_key,cell_value)

    def billing_info_dict_create(self):
        billing_info_matching_key_list = list(self.billing_info_matching_dict.keys())
        billing_info_matching_key_list.append('BILL_MONTH')
        for billing_dict_key in self.billing_dict.keys():
            if 'list' not in billing_dict_key:
                for billing_info_key in self.billing_dict[billing_dict_key].keys():
                    if billing_info_key not in billing_info_matching_key_list:
                        pass
                    else:
                        if billing_info_key == 'CLOUD_NAME':
                            self.billing_info_matching_dict[billing_info_key]['value'] = self.billing_dict[billing_dict_key][billing_info_key] + " CLOUD 외"
                        elif billing_info_key == 'TOTAL_SALES_DISCOUNT_AMT' or billing_info_key == 'TOTAL_COIN_USE_AMT':
                            self.billing_info_matching_dict[billing_info_key]['value'] = -self.billing_dict[billing_dict_key][billing_info_key]
                        elif billing_info_key == 'TOTAL_VAT_INCLUDE_AMT':
                            self.billing_info_matching_dict[billing_info_key]['value'] = self.billing_dict[billing_dict_key][billing_info_key]
                            self.billing_info_matching_dict['TOTAL_VAT_INCLUDE_AMT_2']['value'] = self.billing_dict[billing_dict_key][billing_info_key]

                        elif billing_info_key == 'BILL_MONTH':
                            billing_month=  str(self.billing_dict[billing_dict_key][billing_info_key])
                            doc_num = 'IV'+ billing_month[2:]+str(self.billing_dict['user']['USER_ID']).zfill(4)
                            start_date,end_date,invoice_date,today_date = self.date_format(billing_month)
                            self.billing_info_matching_dict['DOC_NUM']['value'] = doc_num
                            self.billing_info_matching_dict['START_DATE']['value'] = start_date
                            self.billing_info_matching_dict['END_DATE']['value'] = end_date
                            self.billing_info_matching_dict['TODAY_DATE']['value'] = today_date
                            self.billing_info_matching_dict['INVOICE_DATE']['value'] = invoice_date
                        else:
                            self.billing_info_matching_dict[billing_info_key]['value'] = self.billing_dict[billing_dict_key][billing_info_key]
        return self.billing_info_matching_dict
    
    def charge_list_header_create(self):
        self.colleciton_current_row += 1
        self.ws.insert_rows(self.colleciton_current_row)  # 행 추가
        for collection_list_headers_key,collection_list_headers_value in self.collection_list_dict.items():
            self.ws.merge_cells(start_row=self.colleciton_current_row, start_column=collection_list_headers_value['start_column'], end_row=self.colleciton_current_row, end_column=collection_list_headers_value['end_column'])
            cell = self.ws.cell(row=self.colleciton_current_row, column=collection_list_headers_value['start_column'], value=collection_list_headers_key)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(bold=True,size=9)
            cell.fill = self.header_fill
            # 병합된 셀 전체에 테두리 적용
            for col in range(collection_list_headers_value['start_column'], collection_list_headers_value['end_column']+1):
                self.ws.cell(row=self.colleciton_current_row, column=col).border = self.border_style

    def charge_list_footer_create(self):
        # 4. 합계 행 생성
        self.colleciton_current_row += 1
        collection_sum_dict = {
            "서비스명":{'start_column':1,'end_column':18, 'value':'합계(VAT제외)'}, 
            "정상가":{'start_column':19,'end_column':23,'value':self.billing_dict['total_charge_info']['TOTAL_USE_AMT']}, 
            "제공가":{'start_column':24,'end_column':28,'value':self.billing_dict['total_charge_info']['TOTAL_USER_PAY_AMT']}, 
            "비고":{'start_column':29,'end_column':37,'value':''}
            }
        self.ws.insert_rows(self.colleciton_current_row)  # 행 추가
        for collection_sum_key,collection_sum_value in collection_sum_dict.items():
            self.ws.merge_cells(start_row=self.colleciton_current_row, start_column=collection_sum_value['start_column'], end_row=self.colleciton_current_row, end_column=collection_sum_value['end_column'])
            cell = self.ws.cell(row=self.colleciton_current_row, column=collection_sum_value['start_column'], value=collection_sum_value['value'])
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(bold=True,size=9)
            cell.fill = self.header_fill
            if collection_sum_key in ['정상가','제공가']:
                self.charge_amount_format(cell,collection_sum_value['value'])
        for col in range(collection_sum_dict['서비스명']['start_column'],collection_sum_dict['비고']['end_column']+1):
            self.ws.cell(row=self.colleciton_current_row, column=col).border = self.border_style

    def cloud_service_charge_list_create(self):
        cloud_service_charge_list = []
        for service_charge_info in self.billing_dict['cloud_service_charge_list']:
            for type_charge_info in service_charge_info['type_list']:
                charge_info = {
                               'type_name': service_charge_info["CLOUD_SERVICE_CHARGE_NAME"],
                                'product_name': type_charge_info["TYPE_NAME"],
                                'use_amt': type_charge_info["TYPE_USE_AMT"],
                                'pay_amt':0,
                                'notes': type_charge_info["TYPE_NOTES"]}
                cloud_service_charge_list.append(charge_info)
        cloud_service_total_charge_info = {
            'total_use_amt':self.billing_dict['total_cloud_charge_info']['TOTAL_CLOUD_USE_AMT'],
            'total_user_pay_amt':self.billing_dict['total_cloud_charge_info']['TOTAL_CLOUD_USER_PAY_AMT']
            }
        return cloud_service_charge_list,cloud_service_total_charge_info
    
    def third_party_charge_list_create(self):
        third_party_charge_list = []
        for service_charge_info in self.billing_dict['third_party_charge_list']:
            charge_info = {
                            'type_name': service_charge_info["CHARGE_THIRD_PARTY_CATEGORY"],
                            'product_name': service_charge_info["CHARGE_THIRD_PARTY_PRODUCT_NAME"],
                            'use_amt': service_charge_info["CHARGE_THIRD_PARTY_USE_AMT"],
                            'pay_amt':service_charge_info["CHARGE_THIRD_PARTY_USER_PAY_AMT"],
                            'notes': service_charge_info["CHARGE_THIRD_PARTY_NOTES"]}
            third_party_charge_list.append(charge_info)
        empty_charge_info = {'type_name':None,'product_name':None,'use_amt':None,'pay_amt':None,'notes':None}
        third_party_charge_list.append(empty_charge_info)
        third_party_total_charge_info = {
            'total_use_amt':self.billing_dict['total_third_party_charge_info']['TOTAL_CHARGE_THIRD_PARTY_USE_AMT'],
            'total_user_pay_amt':self.billing_dict['total_third_party_charge_info']['TOTAL_CHARGE_THIRD_PARTY_USER_PAY_AMT']
            }
        return third_party_charge_list,third_party_total_charge_info

    def managed_service_charge_list_create(self):
        managed_service_charge_list = []
        for service_charge_info in self.billing_dict['managed_service_charge_list']:
            charge_info = {
                            'type_name': service_charge_info["CHARGE_MANAGED_SERVICE_CATEGORY"],
                            'product_name': service_charge_info["CHARGE_MANAGED_SERVICE_PRODUCT_NAME"],
                            'use_amt': service_charge_info["CHARGE_MANAGED_SERVICE_USE_AMT"],
                            'pay_amt':service_charge_info["CHARGE_MANAGED_SERVICE_USER_PAY_AMT"],
                            'notes': service_charge_info["CHARGE_MANAGED_SERVICE_NOTES"]}
            managed_service_charge_list.append(charge_info)
        empty_charge_info = {'type_name':None,'product_name':None,'use_amt':None,'pay_amt':None,'notes':None}
        managed_service_charge_list.append(empty_charge_info)
        managed_sevice_total_charge_info = {
            'total_use_amt':self.billing_dict['total_managed_service_charge_info']['TOTAL_CHARGE_MANAGED_SERVICE_USE_AMT'],
            'total_user_pay_amt':self.billing_dict['total_managed_service_charge_info']['TOTAL_CHARGE_MANAGED_SERVICE_USER_PAY_AMT']
            }
        return managed_service_charge_list,managed_sevice_total_charge_info
             

    def other_service_charge_list_create(self):
        other_service_charge_list = []
        for service_charge_info in self.billing_dict['other_service_charge_list']:
            charge_info = {
                            'type_name': service_charge_info["CHARGE_OTHER_SERVICE_CATEGORY"],
                            'product_name': service_charge_info["CHARGE_OTHER_SERVICE_PRODUCT_NAME"],
                            'use_amt': service_charge_info["CHARGE_OTHER_SERVICE_USE_AMT"],
                            'pay_amt':service_charge_info["CHARGE_OTHER_SERVICE_USER_PAY_AMT"],
                            'notes': service_charge_info["CHARGE_OTHER_SERVICE_NOTES"]}
            other_service_charge_list.append(charge_info)
        empty_charge_info = {'type_name':None,'product_name':None,'use_amt':None,'pay_amt':None,'notes':None}
        other_service_charge_list.append(empty_charge_info)
        other_sevice_total_charge_info = {
            'total_use_amt':self.billing_dict['total_other_service_charge_info']['TOTAL_CHARGE_OTHER_SERVICE_USE_AMT'],
            'total_user_pay_amt':self.billing_dict['total_other_service_charge_info']['TOTAL_CHARGE_OTHER_SERVICE_USER_PAY_AMT']
            }
        return other_service_charge_list,other_sevice_total_charge_info

    def collection_charge_info_insert(self,collection_charge_info_dict):
        for collection_list_dict_key,collection_list_dict_value in collection_charge_info_dict.items():
            if collection_list_dict_key != '서비스명':
                start_column = collection_list_dict_value['start_column']
                end_column = collection_list_dict_value['end_column']
                cell_key = collection_list_dict_value['key']
                cell_value = collection_list_dict_value['value']
                self.collection_data_insert(start_column, end_column,cell_key, cell_value)

    def collection_charge_list_create(self,service_name,collection_charge_list,collection_total_charge_info):
        # 3. 데이터 삽입 (빈 행 추가)
        start_row = self.colleciton_current_row + 1
        for collection_charge_element in collection_charge_list:
            self.colleciton_current_row += 1
            self.ws.insert_rows(self.colleciton_current_row)
            
            self.collection_list_dict['청구항목']['value'] = collection_charge_element["type_name"]
            self.collection_list_dict['상품명']['value'] = collection_charge_element["product_name"]
            self.collection_list_dict['정상가']['value'] = collection_charge_element["use_amt"]
            self.collection_list_dict['제공가']['value'] = collection_charge_element["pay_amt"]
            self.collection_list_dict['비고']['value'] = collection_charge_element["notes"]
            self.collection_charge_info_insert(self.collection_list_dict)

        self.colleciton_current_row += 1
        self.ws.insert_rows(self.colleciton_current_row)
        self.collection_middle_sum_dict['소계']['value'] = '소계'
        self.collection_middle_sum_dict['정상가']['value'] = collection_total_charge_info['total_use_amt']
        self.collection_middle_sum_dict['제공가']['value'] = collection_total_charge_info['total_user_pay_amt']
        self.collection_middle_sum_dict['비고']['value'] = ''
            
        self.collection_charge_info_insert(self.collection_middle_sum_dict)
                    
        self.ws.merge_cells(start_row=start_row, start_column=1, end_row=self.colleciton_current_row, end_column=4)
        cell = self.ws.cell(row=start_row, column=1, value=service_name)
        cell.alignment = self.center_alignment
        cell.font = Font(size=9)
        
        for row in range(start_row, self.colleciton_current_row+1):
            for col in range(1, self.collection_list_dict['비고']['end_column']+1):
                self.ws.cell(row=row, column=col).border = self.border_style
    
    def charge_summary_create(self):    
        for billing_info_matching_key,billing_info_matching_value in self.billing_info_matching_dict.items():
            cell_row = billing_info_matching_value['row']
            cell_column = billing_info_matching_value['column']
            cell_value = billing_info_matching_value['value']
            cell_key = billing_info_matching_key
            cell = self.ws.cell(row=cell_row, column=cell_column, value=cell_value)
            self.cell_format_setting(cell,cell_key,cell_value)
    
    
    def invoice_create(self):
        self.billing_info_dict_create()

        # 모든 셀에 중앙 정렬 적용
        for row in self.ws.iter_rows():  # 워크시트의 모든 셀 반복
            self.ws.row_dimensions[row[0].row].height = 16.5  # 행 높이를 16.5로 설정
            for cell in row:
                if cell.row not in [16,30,33,35,36,37,38]:
                    cell.alignment = self.center_alignment

        self.image_create()

        self.charge_summary_create()

        self.charge_list_header_create()
        
        cloud_service_charge_list,cloud_service_total_charge_info = self.cloud_service_charge_list_create()
        service_name = self.billing_dict['user']['CLOUD_NAME']+" CLOUD"
        self.collection_charge_list_create(service_name,cloud_service_charge_list,cloud_service_total_charge_info)
        
        third_party_charge_list,third_party_total_charge_info = self.third_party_charge_list_create()
        service_name = '3rd party S/W'
        self.collection_charge_list_create(service_name,third_party_charge_list,third_party_total_charge_info)
        
        managed_service_charge_list,managed_sevice_total_charge_info = self.managed_service_charge_list_create()
        service_name = '매니지드 서비스'
        self.collection_charge_list_create(service_name,managed_service_charge_list,managed_sevice_total_charge_info)
        
        other_service_charge_list,other_sevice_total_charge_info = self.other_service_charge_list_create()
        service_name = '기타'
        self.collection_charge_list_create(service_name,other_service_charge_list,other_sevice_total_charge_info)

        self.charge_list_footer_create()
        # 3. 수정된 파일 저장
        file_path = "bemon_manage_func/invoice_result.xlsx"
        self.wb.save(file_path)
        return file_path